from airflow import DAG
from datetime import timedelta, datetime
from airflow.utils.dates import days_ago
from airflow.operators.bash import BashOperator
from airflow.operators.python import PythonOperator
from airflow.operators.email import EmailOperator
import requests
from requests import get
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import openpyxl
from datetime import date
from datetime import timedelta
from datetime import datetime
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import xlsxwriter
import re
import pandas as pd
import os
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import sys
import json
import pytz
#print(sys.setrecursionlimit(4000))

chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument('--disable-dev-shm-usage')
options = Options()

options.headless = True

def date_range(start, end):
    delta = end - start
    days = [start + timedelta(days=i) for i in range(delta.days + 1)]
    return days

def get_data(date):
    try:
        url = 'https://archives.nseindia.com/content/nsccl/fao_participant_oi_'+date.strftime("%d%m%Y")+'.csv'
        headers = {"Accept-Language": "en-US, en;q=0.5"}
        r = requests.get(url, headers=headers,timeout=3).status_code
        df = pd.read_csv(url)
        titles  = [] 
        for i in range(len(df.iloc[0])):
            titles.append(df.iloc[0][i])
        df.columns = titles
        df.drop(index=df.index[0], axis=0, inplace=True) 
        df = df[['Client Type','Option Index Call Long','Option Index Put Long','Option Index Call Short','Option Index Put Short']]
        df.set_index(df.columns[0],inplace=True)
        values = df.loc['TOTAL',:].tolist()
        return date,values
    except requests.exceptions.ReadTimeout:
        return None,None
    except:
        return None,None

def extract(**kwargs):
    start = datetime(2020,2,14)
    end = datetime.now()
    dates = date_range(start, end)
    d = []
    v = []
    for date in dates:
        print(date)
        dd,values = get_data(date)
        if dd == None and values == None:
            continue
        else:
            d.append(dd)
            v.append(values)
    
    return d,v

def transform(**context):
    dates = list(context['ti'].xcom_pull(task_ids='extract')[0])
    values = list(context['ti'].xcom_pull(task_ids='extract')[1]) 
    date = []
    for i in dates:
        x = i.strftime('%d-%m-%Y')
        date.append(x)

    return date,values

def load(**context):
    date = list(context['ti'].xcom_pull(task_ids='transform')[0])
    values = list(context['ti'].xcom_pull(task_ids='transform')[1])
    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    check_date = "14-02-2020"
    check_date = datetime.strptime(check_date, "%d-%m-%Y")
    sheet = wb.worksheets[1]
    sheet_a =[]
    for cell in sheet["A"]:
        sheet_a.append(cell.value)
    for i in range(len(date)):
        x = date[i]
        p = values[i] 
        y = datetime.strptime(x, "%d-%m-%Y")
        if y.date()<check_date.date():
            continue
        if x in sheet_a:
                for cell in sheet['A']:
                    if cell.value == x:
                        sheet["BC" + str(cell.row)] = float(p[0])
                        sheet["BD" + str(cell.row)] = float(p[1])
                        sheet["BE" + str(cell.row)] = float(p[2])
                        sheet["BF" + str(cell.row)] = float(p[3])

                        break
        else:
            row = sheet.max_row+1
            sheet["A" + str(row)] = x
            sheet["BC" + str(row)] = float(p[0])
            sheet["BD" + str(row)] = float(p[1])
            sheet["BE" + str(row)] = float(p[2])
            sheet["BF" + str(row)] = float(p[3])
            sheet_a.append(x)
    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')


dt = datetime.now()

default_args = {
    'owner': 'airflow',
    'depends_on_past': False,
    'start_date': datetime(2023,1,12),
    'email': ['nikheleshbhattacharyya@gmail.com'],
    'email_on_failure': True,
    'email_on_retry': False,
    'retries': 3,
    'retry_delay': timedelta(minutes=2)
}

schedule_interval = '0 */2 * * *'

dag = DAG(
    dag_id = 'NSE_Open_Interest',
    default_args = default_args,
    schedule = '@daily'
)

task_start = BashOperator(
    task_id = 'start_task',
    bash_command = 'echo start',
    dag = dag
)

task1 = PythonOperator(
    task_id = 'extract',
    python_callable = extract,
    provide_context = True,
    dag = dag
)

task2 = PythonOperator(
    task_id = 'transform',
    python_callable = transform,
    provide_context = True,
    dag = dag
)

task3 = PythonOperator(
    task_id = 'load',
    python_callable = load,
    provide_context = True,
    dag = dag
)


#send_email = EmailOperator(
#    task_id = 'send_email',
#    to = ['nikheleshbhattacharyya@gmail.com'],
#    subject = 'FX_Project data ',
#    html_content = '''
#    {date} Data data has been scraped!
#    '''.format(date = datetime.now()),
#    dag = dag
#)

finish_start = BashOperator(
    task_id = 'finish_task',
    bash_command = 'echo finish',
    dag = dag
)

task_start>>task1>>task2>>task3>>finish_start