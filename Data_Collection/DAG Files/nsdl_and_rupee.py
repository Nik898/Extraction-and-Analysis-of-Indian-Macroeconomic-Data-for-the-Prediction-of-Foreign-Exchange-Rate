from airflow import DAG
from datetime import timedelta, datetime
from airflow.utils.dates import days_ago
from airflow.operators.bash import BashOperator
from airflow.operators.python import PythonOperator
from airflow.operators.email import EmailOperator
import requests
from requests import get
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import openpyxl
from datetime import date
from datetime import timedelta
from datetime import datetime
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import xlsxwriter
import re
import pandas as pd
import os
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import sys
import json
import pytz
from fake_useragent import UserAgent
from dateutil.relativedelta import relativedelta
from forex_python.converter import CurrencyRates
#print(sys.setrecursionlimit(4000))

chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument('--disable-dev-shm-usage')
options = Options()

options.headless = True

def nthOccurence(temp, n):
    val = -1
    for i in range(0, n): val = temp.find('<td', val + 1)
    temp = temp[val+4:]
    temp = temp[:temp.find("</td>")]
    temp = temp.replace("(","-")
    temp = temp.replace(")", "")
    return temp

def date_find(temp, n):
    val = -1
    for i in range(0, n): val = temp.find('<td', val + 1)
    temp = temp[val+17:]
    temp = temp[:temp.find("</td>")]
    temp = temp.replace("(","-")
    temp = temp.replace(")", "")
    return temp

def get_dates():
    date_arr = pd.date_range('2020-01-14',str(datetime.now()), 
                freq='MS').strftime("%d-%b-%Y").tolist()
    dates_arr = []
    for date in date_arr:
        date = datetime.strptime(date,"%d-%b-%Y")
        date = date + relativedelta(day=31)
        date = date.strftime("%d-%b-%Y")
        dates_arr.append(date)
    return dates_arr

def get_data(driver,dates_arr):
    main_arr = []
    for date in dates_arr:
        try:
            driver.execute_script('document.getElementById("txtDate").setAttribute("disabled", false)')
            driver.execute_script("document.getElementById('txtDate').setAttribute('value','"+ date+ "')")
            driver.execute_script("document.getElementById('hdnDate').setAttribute('value','"+ date+ "')")
            driver.execute_script("document.getElementById('btnSubmit1').click()")

            text = driver.find_element(By.ID,"dvArchiveData")
            table = text.find_elements(By.TAG_NAME,"table")[0].get_attribute("innerHTML")
            
            name_arr = table.split("<tr")
            noOfDays = (len(name_arr)-30)/13

            for a in range(int(noOfDays)):
                i = 13*a
                FPI_arr = []
                temp = date_find(name_arr[3+i],1)
                #FPI_arr.append(datetime.strptime(temp, '%d-%b-%Y'))
                x = datetime.strptime(temp, '%d-%b-%Y')
                check_date = "14-02-2020"
                check_date = datetime.strptime(check_date, "%d-%m-%Y")
                if check_date.date() > x.date():
                    print(x)
                    continue
                else:
                    print(x)
                    xx = x.strftime('%d-%m-%Y')
                    FPI_arr.append(xx)
                    FPI_arr.append(float(nthOccurence(str(name_arr[3+i]),7)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[4+i]),5)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[5+i]),5)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[6+i]),6)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[7+i]),5)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[8+i]),5)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[9+i]),6)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[10+i]),5)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[11+i]),5)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[12+i]),6)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[13+i]),5)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[14+i]),5)))
                    FPI_arr.append(float(nthOccurence(str(name_arr[15+i]),5)))
                    main_arr.append(FPI_arr)
        except:
            continue
    return main_arr

def get_nsdl_fpi(driver):
    dates = get_dates()
    main_arr = get_data(driver,dates)
    return main_arr   

def extract_nsdl_fpi(**kwargs):
    ua = UserAgent()
    userAgent = ua.random
    url = "https://www.fpi.nsdl.co.in"
    driver = webdriver.Chrome('/usr/local/bin/chromedriver2',options=chrome_options)
    driver.execute_cdp_cmd('Network.setUserAgentOverride', {"userAgent": userAgent})
    driver.implicitly_wait(30)
    driver.get(url)
    head = driver.find_element(By.ID,"HeaderControl_dvHlinks")
    btn = head.find_elements(By.CLASS_NAME, "tabslist")[1].click()
    archive_tag = driver.find_elements(By.TAG_NAME, "li")[4].click()
    arr = get_nsdl_fpi(driver)
    driver.quit()
    return arr

def load_nsdl_fpi(**context):
    arr = list(context['ti'].xcom_pull(task_ids='extract_data'))
    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    sheet = wb.worksheets[1]
    sheet_a =[]
    for cell in sheet["A"]:
        sheet_a.append(cell.value)
    for i in range(len(arr)):
        x = arr[i][0]
        p = arr[i][1:] 
        if x in sheet_a:
                for cell in sheet['A']:
                    if cell.value == x:
                        sheet["BG" + str(cell.row)] = float(p[0])
                        sheet["BH" + str(cell.row)] = float(p[1])
                        sheet["BI" + str(cell.row)] = float(p[2])
                        sheet["BJ" + str(cell.row)] = float(p[3])
                        sheet["BK" + str(cell.row)] = float(p[4])
                        sheet["BL" + str(cell.row)] = float(p[5])
                        sheet["BM" + str(cell.row)] = float(p[6])
                        sheet["BN" + str(cell.row)] = float(p[7])
                        sheet["BO" + str(cell.row)] = float(p[8])
                        sheet["BP" + str(cell.row)] = float(p[9])
                        sheet["BQ" + str(cell.row)] = float(p[10])
                        sheet["BR" + str(cell.row)] = float(p[11])
                        sheet["BS" + str(cell.row)] = float(p[12])
                        break
        else:
            row = sheet.max_row+1
            sheet["A" + str(row)] = x
            sheet["BG" + str(cell.row)] = float(p[0])
            sheet["BH" + str(cell.row)] = float(p[1])
            sheet["BI" + str(cell.row)] = float(p[2])
            sheet["BJ" + str(cell.row)] = float(p[3])
            sheet["BK" + str(cell.row)] = float(p[4])
            sheet["BL" + str(cell.row)] = float(p[5])
            sheet["BM" + str(cell.row)] = float(p[6])
            sheet["BN" + str(cell.row)] = float(p[7])
            sheet["BO" + str(cell.row)] = float(p[8])
            sheet["BP" + str(cell.row)] = float(p[9])
            sheet["BQ" + str(cell.row)] = float(p[10])
            sheet["BR" + str(cell.row)] = float(p[11])
            sheet["BS" + str(cell.row)] = float(p[12])
            sheet_a.append(x)
    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')

def extract_rupee(**kwargs):
    c = CurrencyRates()
    usdtoinr =[]
    dates =[]
    start = datetime(2020,2,14)
    end = datetime.now()
    delta = end - start
    days = [start + timedelta(days=i) for i in range(delta.days + 1)]
    for d in days:
        print(d)
        dates.append(d.strftime("%d-%m-%Y"))
        value = c.get_rate('USD', 'INR',d)
        value = round(value, 2)
        usdtoinr.append(value)

    return dates,usdtoinr

def load_rupee(**context):
    dates = list(context['ti'].xcom_pull(task_ids='extract_data2')[0])
    values = list(context['ti'].xcom_pull(task_ids='extract_data2')[1])
    wb = load_workbook(filename='/root/airflow/files/Data.xlsx')
    sheet = wb.worksheets[1]
    sheet_a =[]
    for cell in sheet["A"]:
        sheet_a.append(cell.value)
    for i in range(len(dates)):
        x = dates[i]
        p = values[i]
        if x in sheet_a:
                for cell in sheet['A']:
                    if cell.value == x:
                        sheet["BT" + str(cell.row)] = float(p)
                        break
        else:
            row = sheet.max_row+1
            sheet["A" + str(row)] = x
            sheet["BT" + str(cell.row)] = float(p)
            sheet_a.append(x)
    wb.close()
    wb.save(filename='/root/airflow/files/Data.xlsx')

dt = datetime.now()

default_args = {
    'owner': 'airflow',
    'depends_on_past': False,
    'start_date': datetime(2023,2,23),
    'email': ['nikheleshbhattacharyya@gmail.com'],
    'email_on_failure': True,
    'email_on_retry': False,
    'retries': 3,
    'retry_delay': timedelta(minutes=2)
}

schedule_interval = '0 */2 * * *'

dag = DAG(
    dag_id = 'nsdlfpi_and_rupee',
    default_args = default_args,
    schedule = '@daily'
)

task_start = BashOperator(
    task_id = 'start_task',
    bash_command = 'echo start',
    dag = dag
)

task1 = PythonOperator(
    task_id = 'extract_data',
    python_callable = extract_nsdl_fpi,
    provide_context = True,
    dag = dag
)

task2 = PythonOperator(
    task_id = 'load_data',
    python_callable = load_nsdl_fpi,
    provide_context = True,
    dag = dag
)

task3 = PythonOperator(
    task_id = 'extract_data2',
    python_callable = extract_rupee,
    provide_context = True,
    dag = dag
)

task4 = PythonOperator(
    task_id = 'load_data2',
    python_callable = load_rupee,
    provide_context = True,
    dag = dag
)

send_email = EmailOperator(
    task_id = 'send_email',
    to = ['nikheleshbhattacharyya@gmail.com'],
    subject = 'FX_Data ',
    html_content = '''
    {date} Data data has been scraped!
    '''.format(date = datetime.now()),
    files = ['/root/airflow/files/Data.xlsx'],
    dag = dag
)

finish_start = BashOperator(
    task_id = 'finish_task',
    bash_command = 'echo finish',
    dag = dag
)

task_start>>task1>>task2>>task3>>task4>>send_email>>finish_start


